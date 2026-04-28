import yfinance as yf
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

COMPS = {
    "Apple":     "AAPL",
    "Microsoft": "MSFT",
    "Alphabet":  "GOOGL",
    "Meta":      "META",
    "Amazon":    "AMZN",
}

def fetch_comps_data(tickers):
    rows = []
    for name, ticker in tickers.items():
        print(f"  Fetching {name} ({ticker})...")
        try:
            t = yf.Ticker(ticker)
            info = t.info
            market_cap = info.get("marketCap", None)
            total_debt = info.get("totalDebt", 0) or 0
            cash       = info.get("totalCash", 0) or 0
            ebitda     = info.get("ebitda", None)
            revenue    = info.get("totalRevenue", None)
            net_income = info.get("netIncomeToCommon", None)
            price      = info.get("currentPrice", None)
            ev = (market_cap + total_debt - cash) if market_cap else None
            ev_ebitda  = round(ev / ebitda, 1)            if ev and ebitda                          else "N/A"
            ev_revenue = round(ev / revenue, 1)           if ev and revenue                         else "N/A"
            pe_ratio   = round(market_cap / net_income, 1) if market_cap and net_income and net_income > 0 else "N/A"
            rows.append({
                "Company":        name,
                "Ticker":         ticker,
                "Market Cap ($B)": round(market_cap / 1e9, 1) if market_cap else "N/A",
                "EV ($B)":         round(ev / 1e9, 1)         if ev         else "N/A",
                "Revenue ($B)":    round(revenue / 1e9, 1)    if revenue    else "N/A",
                "EBITDA ($B)":     round(ebitda / 1e9, 1)     if ebitda     else "N/A",
                "EV/EBITDA (x)":   ev_ebitda,
                "EV/Revenue (x)":  ev_revenue,
                "P/E (x)":         pe_ratio,
                "Price ($)":       round(price, 2)            if price      else "N/A",
            })
        except Exception as e:
            print(f"    Warning: {ticker}: {e}")
            rows.append({"Company": name, "Ticker": ticker})
    return pd.DataFrame(rows)

def add_summary_stats(df):
    numeric_cols = ["EV/EBITDA (x)", "EV/Revenue (x)", "P/E (x)"]
    summary_rows = []
    for label in ["Mean", "Median", "Min", "Max"]:
        row = {"Company": label, "Ticker": "—"}
        for col in numeric_cols:
            vals = pd.to_numeric(df[col], errors="coerce").dropna()
            if len(vals) == 0:
                row[col] = "N/A"
            elif label == "Mean":   row[col] = round(vals.mean(), 1)
            elif label == "Median": row[col] = round(vals.median(), 1)
            elif label == "Min":    row[col] = round(vals.min(), 1)
            elif label == "Max":    row[col] = round(vals.max(), 1)
        summary_rows.append(row)
    return pd.concat([df, pd.DataFrame(summary_rows)], ignore_index=True)

def write_to_excel(df, excel_path):
    wb = load_workbook(excel_path)
    if "Comps" in wb.sheetnames:
        del wb["Comps"]
    ws = wb.create_sheet("Comps")
    header_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="1F3864")
    summary_fill = PatternFill("solid", fgColor="D9E1F2")
    center       = Alignment(horizontal="center", vertical="center")
    left         = Alignment(horizontal="left",   vertical="center")
    border       = Border(bottom=Side(style="thin", color="BFBFBF"),
                          right =Side(style="thin", color="BFBFBF"))
    ws.merge_cells("A1:J1")
    tc = ws["A1"]
    tc.value     = f"Comparable Company Analysis  |  Pulled: {datetime.now().strftime('%d %b %Y, %H:%M')}"
    tc.font      = Font(name="Calibri", bold=True, size=13, color="1F3864")
    tc.alignment = left
    ws.row_dimensions[1].height = 22
    headers = list(df.columns)
    ws.row_dimensions[3].height = 18
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center; cell.border = border
    summary_labels = {"Mean", "Median", "Min", "Max"}
    for ri, row_data in df.iterrows():
        er = 4 + ri
        is_sum = row_data.get("Company") in summary_labels
        ws.row_dimensions[er].height = 16
        for ci, value in enumerate(row_data, 1):
            cell = ws.cell(row=er, column=ci, value=value)
            cell.border = border
            cell.alignment = center if ci > 2 else left
            if is_sum:
                cell.fill = summary_fill
                cell.font = Font(name="Calibri", bold=True, size=10, color="1F3864")
            else:
                cell.font = Font(name="Calibri", size=10)
    for i, w in enumerate([18,9,16,14,14,13,16,16,12,12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    wb.save(excel_path)
    print(f"\n  Comps tab written to: {excel_path}")

if __name__ == "__main__":
    EXCEL_PATH = os.path.expanduser("~/dcf-valuation-engine/model/DCF_Valuation_Engine.xlsx")
    print("\nDCF Valuation Engine — Fetching Comparable Companies\n" + "-" * 55)
    df = fetch_comps_data(COMPS)
    df = add_summary_stats(df)
    print("\n-- Comps Table --")
    print(df[["Company", "EV/EBITDA (x)", "EV/Revenue (x)", "P/E (x)"]].to_string(index=False))
    print("\n-- Writing to Excel --")
    write_to_excel(df, EXCEL_PATH)
    print("\nDone. Open DCF_Valuation_Engine.xlsx and check the Comps tab.\n")
